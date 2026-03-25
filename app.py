from io import BytesIO
from math import ceil

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from code_data import (
    OCCUPANT_LOAD_STANDARDS,
    PLUMBING_FIXTURE_CATEGORIES,
    SEPARATE_FACILITIES_RULES,
)

st.set_page_config(page_title="CodeCalc", layout="wide")

# ---------------------------------------------------------------------------
# Global styles
# ---------------------------------------------------------------------------

st.markdown("""
<style>
    /* tighten overall padding */
    .block-container { padding-top: 1.5rem; padding-bottom: 1rem; }

    /* section headers */
    h2 { margin-top: 1.2rem !important; margin-bottom: 0.3rem !important; }
    h3 { margin-top: 1rem !important; margin-bottom: 0.2rem !important; }

    /* metric cards — compact */
    [data-testid="stMetric"] {
        background: #f8f9fb;
        border: 1px solid #e0e4ea;
        border-radius: 6px;
        padding: 0.55rem 0.7rem;
        text-align: center;
    }
    [data-testid="stMetricLabel"] {
        font-size: 0.72rem !important;
        font-weight: 600;
        color: #5a6577;
        text-transform: uppercase;
        letter-spacing: 0.03em;
    }
    [data-testid="stMetricValue"] {
        font-size: 1.35rem !important;
        font-weight: 700;
        color: #1a1f2b;
    }

    /* totals row — accent background */
    .totals-card [data-testid="stMetric"] {
        background: #e8f0fe;
        border-color: #a8c7fa;
    }
    .totals-card [data-testid="stMetricValue"] { color: #1a56db; }

    /* plumbing category block */
    .plumbing-block {
        background: #ffffff;
        border: 1px solid #e0e4ea;
        border-radius: 8px;
        padding: 1rem 1.2rem 0.6rem;
        margin-bottom: 0.8rem;
    }

    /* separate‐facilities badges */
    .sep-required {
        display: inline-block;
        background: #fef3c7;
        border: 1px solid #f59e0b;
        border-radius: 4px;
        padding: 0.25rem 0.6rem;
        font-size: 0.78rem;
        color: #92400e;
        margin-top: 0.4rem;
    }
    .sep-not-required {
        display: inline-block;
        background: #ecfdf5;
        border: 1px solid #34d399;
        border-radius: 4px;
        padding: 0.25rem 0.6rem;
        font-size: 0.78rem;
        color: #065f46;
        margin-top: 0.4rem;
    }

    /* data editor — give it breathing room */
    [data-testid="stDataFrame"] { margin-bottom: 0.6rem; }

    /* sidebar polish */
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p {
        font-size: 0.85rem;
    }
</style>
""", unsafe_allow_html=True)

DEFAULT_STANDARD = "IBC 2021"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get_standard_table(standard_name: str):
    return OCCUPANT_LOAD_STANDARDS[standard_name]


def get_plumbing_category_table():
    return PLUMBING_FIXTURE_CATEGORIES


def format_function_label(item: dict) -> str:
    return f"{item['name'].replace(', gross', '').replace(', net', '')}, {int(item['factor'])} {item['area_type']}"


def build_lookup(standard_name: str):
    return {format_function_label(item): item for item in get_standard_table(standard_name)}


def format_plumbing_category_label(item: dict) -> str:
    return item["label"]


def calculate_occupant_load(area: float, factor: float) -> int:
    if factor <= 0:
        return 0
    return ceil(area / factor)


def calculate_fixture_raw(occupants: int, ratio) -> float:
    """Return the exact fractional fixture count (no rounding)."""
    if ratio is None or ratio == 0:
        return 0.0
    if isinstance(ratio, dict):
        if "fixed" in ratio:
            return float(ratio["fixed"])
        first_block = min(occupants, ratio["first"])
        remainder = max(occupants - ratio["first"], 0)
        count = 0.0
        if first_block > 0:
            count += first_block / ratio["per"]
        if remainder > 0:
            count += remainder / ratio["then_per"]
        return count
    if isinstance(ratio, (int, float)) and ratio > 0:
        return occupants / ratio
    return 0.0


def format_ratio_text(ratio) -> str:
    if ratio is None:
        return "\u2014"
    if isinstance(ratio, dict):
        if "fixed" in ratio:
            return str(ratio["fixed"])
        return f"1 per {ratio['per']} (first {ratio['first']}), then 1 per {ratio['then_per']}"
    if isinstance(ratio, (int, float)) and ratio > 0:
        return f"1 per {int(ratio)}"
    return "\u2014"


def fmt_frac(val: float) -> str:
    """Format a fractional fixture value for display: 2 decimals if fractional, integer if whole."""
    if val == 0:
        return "\u2014"
    if val == int(val):
        return str(int(val))
    return f"{val:.2f}"


def split_occupants(total_occupants: int, male_ratio: float) -> tuple[int, int]:
    male = ceil(total_occupants * male_ratio)
    female = max(total_occupants - male, 0)
    return male, female


def evaluate_separate_facilities(category_item: dict, occupants: int) -> str:
    rule_key = category_item.get("separate_facilities_rule", "general")
    rule = SEPARATE_FACILITIES_RULES.get(rule_key, SEPARATE_FACILITIES_RULES["general"])

    if not rule["required"]:
        if rule["exceptions"]:
            return rule["exceptions"][0]["text"]
        return "Not required"

    if occupants <= 15:
        return "Not required \u2014 Exception 2: Total occupant load is 15 or fewer (IBC B2902.2)"

    for exc in rule["exceptions"]:
        if "max_occupants" in exc and occupants <= exc["max_occupants"]:
            return f"Not required \u2014 {exc['text']}"

    return "Required \u2014 Separate facilities shall be provided for each sex (IBC B2902.2)"


# ---------------------------------------------------------------------------
# Room schedule (Section 1004 occupant load)
# ---------------------------------------------------------------------------

def starter_rows():
    return pd.DataFrame([
        {"Story": "Level 1", "Room Number": "101", "Room Name": "Lobby", "Room Area": 1200.0,
         "Room Function": "Business areas, 150 gross"},
        {"Story": "Level 1", "Room Number": "102", "Room Name": "Conference", "Room Area": 450.0,
         "Room Function": "Assembly without fixed seats, unconcentrated (tables and chairs), 15 net"},
    ])


def enrich_rows(df: pd.DataFrame, standard_name: str) -> pd.DataFrame:
    lookup = build_lookup(standard_name)
    rows = []
    for _, row in df.iterrows():
        selected = lookup.get(row.get("Room Function"))
        if not selected:
            continue
        area = float(row.get("Room Area") or 0)
        factor = float(selected["factor"])
        occupants = calculate_occupant_load(area, factor)
        rows.append({
            "Story": row.get("Story", ""),
            "Room Number": row.get("Room Number", ""),
            "Room Name": row.get("Room Name", ""),
            "Room Area": area,
            "Room Function": format_function_label(selected),
            "Occupants": occupants,
        })
    return pd.DataFrame(rows)


def build_schedule_lines(schedule_df: pd.DataFrame) -> pd.DataFrame:
    if schedule_df.empty:
        return pd.DataFrame(columns=["Story", "Room Number", "Room Name", "Room Area", "Room Function", "Occupants"])
    line_rows = []
    for story, group in schedule_df.groupby("Story", sort=False):
        story_total = int(group["Occupants"].sum())
        for _, row in group.iterrows():
            line_rows.append({k: row[k] for k in ["Story", "Room Number", "Room Name", "Room Area", "Room Function", "Occupants"]})
        line_rows.append({"Story": story, "Room Number": "", "Room Name": "", "Room Area": None,
                          "Room Function": "Story Total Occupants", "Occupants": story_total})
    return pd.DataFrame(line_rows)


# ---------------------------------------------------------------------------
# Plumbing fixture analysis (Table B2902.1)
# ---------------------------------------------------------------------------

def plumbing_starter_rows():
    return pd.DataFrame([{"Plumbing Category": "2 - Business", "Occupants": 8}])


def build_plumbing_analysis(plumbing_df: pd.DataFrame, male_ratio: float) -> list[dict]:
    """Compute raw fractional fixture counts per category row. Rounding happens at the totals only."""
    if plumbing_df.empty:
        return []

    plumbing_by_label = {format_plumbing_category_label(item): item for item in get_plumbing_category_table()}
    results = []
    for _, row in plumbing_df.iterrows():
        category_label = row.get("Plumbing Category")
        category_item = plumbing_by_label.get(category_label)
        if not category_item:
            continue
        occupants = int(row.get("Occupants") or 0)
        if occupants <= 0:
            continue
        male_occ, female_occ = split_occupants(occupants, male_ratio)

        results.append({
            "classification_no": category_item["classification_no"],
            "classification": category_item["classification"],
            "description": category_item["description"],
            "label": category_label,
            "occupants": occupants,
            "male_occ": male_occ,
            "female_occ": female_occ,
            "wc_male": calculate_fixture_raw(male_occ, category_item["wc_male"]),
            "wc_female": calculate_fixture_raw(female_occ, category_item["wc_female"]),
            "lav_male": calculate_fixture_raw(male_occ, category_item["lav_male"]),
            "lav_female": calculate_fixture_raw(female_occ, category_item["lav_female"]),
            "bath": calculate_fixture_raw(occupants, category_item["bath"]),
            "drinking": calculate_fixture_raw(occupants, category_item["drinking"]),
            "service_sink": calculate_fixture_raw(occupants, category_item["service_sink"]),
            "other": category_item["other"] if category_item["other"] else "\u2014",
            "separate_facilities": evaluate_separate_facilities(category_item, occupants),
            "code_ref": category_item["code_ref"],
            "wc_male_ratio": category_item["wc_male"],
            "wc_female_ratio": category_item["wc_female"],
            "lav_male_ratio": category_item["lav_male"],
            "lav_female_ratio": category_item["lav_female"],
            "bath_ratio": category_item["bath"],
            "drinking_ratio": category_item["drinking"],
            "service_sink_ratio": category_item["service_sink"],
        })

    return results


FIXTURE_KEYS = ["wc_male", "wc_female", "lav_male", "lav_female", "bath", "drinking", "service_sink"]


def sum_fixtures(analysis: list[dict]) -> dict:
    """Sum raw fractional fixture values across all categories."""
    sums = {k: 0.0 for k in FIXTURE_KEYS}
    for item in analysis:
        for k in FIXTURE_KEYS:
            sums[k] += item[k]
    sums["occupants"] = sum(r["occupants"] for r in analysis)
    sums["male_occ"] = sum(r["male_occ"] for r in analysis)
    sums["female_occ"] = sum(r["female_occ"] for r in analysis)
    return sums


def ceil_fixtures(sums: dict) -> dict:
    """Round up the accumulated fractional totals to get final required counts."""
    return {k: ceil(sums[k]) if sums[k] > 0 else 0 for k in FIXTURE_KEYS}


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def build_plumbing_export_df(analysis: list[dict]) -> pd.DataFrame:
    columns = [
        "No.", "Classification", "Description",
        "Occupants", "Male Occ", "Female Occ",
        "WC (M)", "WC (F)", "Lav (M)", "Lav (F)",
        "Bathtubs/Showers", "Drinking Fountains", "Service Sinks",
        "Separate Facilities", "Code Basis",
    ]
    if not analysis:
        return pd.DataFrame(columns=columns)

    sums = sum_fixtures(analysis)
    required = ceil_fixtures(sums)

    rows = []
    for item in analysis:
        rows.append({
            "No.": item["classification_no"],
            "Classification": item["classification"],
            "Description": item["description"],
            "Occupants": item["occupants"],
            "Male Occ": item["male_occ"],
            "Female Occ": item["female_occ"],
            "WC (M)": round(item["wc_male"], 2),
            "WC (F)": round(item["wc_female"], 2),
            "Lav (M)": round(item["lav_male"], 2),
            "Lav (F)": round(item["lav_female"], 2),
            "Bathtubs/Showers": round(item["bath"], 2) if item["bath"] > 0 else "\u2014",
            "Drinking Fountains": round(item["drinking"], 2),
            "Service Sinks": round(item["service_sink"], 2) if item["service_sink"] > 0 else "\u2014",
            "Separate Facilities": item["separate_facilities"],
            "Code Basis": item["code_ref"],
        })

    # Running accumulation row
    rows.append({
        "No.": "",
        "Classification": "ACCUMULATED (raw)",
        "Description": "",
        "Occupants": sums["occupants"],
        "Male Occ": sums["male_occ"],
        "Female Occ": sums["female_occ"],
        "WC (M)": round(sums["wc_male"], 2),
        "WC (F)": round(sums["wc_female"], 2),
        "Lav (M)": round(sums["lav_male"], 2),
        "Lav (F)": round(sums["lav_female"], 2),
        "Bathtubs/Showers": round(sums["bath"], 2) if sums["bath"] > 0 else "\u2014",
        "Drinking Fountains": round(sums["drinking"], 2),
        "Service Sinks": round(sums["service_sink"], 2) if sums["service_sink"] > 0 else "\u2014",
        "Separate Facilities": "",
        "Code Basis": "",
    })

    # Final required (rounded up)
    rows.append({
        "No.": "",
        "Classification": "TOTAL REQUIRED",
        "Description": "",
        "Occupants": sums["occupants"],
        "Male Occ": sums["male_occ"],
        "Female Occ": sums["female_occ"],
        "WC (M)": required["wc_male"],
        "WC (F)": required["wc_female"],
        "Lav (M)": required["lav_male"],
        "Lav (F)": required["lav_female"],
        "Bathtubs/Showers": required["bath"] if required["bath"] > 0 else "\u2014",
        "Drinking Fountains": required["drinking"],
        "Service Sinks": required["service_sink"] if required["service_sink"] > 0 else "\u2014",
        "Separate Facilities": "",
        "Code Basis": "IBC 2021 Table B2902.1",
    })

    return pd.DataFrame(rows)


def build_workbook(schedule_df: pd.DataFrame, standard_name: str, plumbing_analysis: list[dict], male_ratio: float) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Occupant Load Schedule"

    export_df = build_schedule_lines(schedule_df)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    accent_fill = PatternFill("solid", fgColor="E8F0FE")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = list(export_df.columns)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in export_df.itertuples(index=False, name=None):
        ws.append(list(row))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=5).value == "Story Total Occupants":
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="F2F2F2")

    for col_idx, w in {1: 14, 2: 14, 3: 28, 4: 14, 5: 42, 6: 14}.items():
        ws.column_dimensions[chr(64 + col_idx)].width = w

    # Plumbing sheet
    plumbing_export = build_plumbing_export_df(plumbing_analysis)
    plumbing_ws = wb.create_sheet("Plumbing Fixtures")
    plumbing_headers = list(plumbing_export.columns)
    plumbing_ws.append(plumbing_headers)
    for cell in plumbing_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in plumbing_export.itertuples(index=False, name=None):
        plumbing_ws.append(list(row))
    for row in plumbing_ws.iter_rows(min_row=2, max_row=plumbing_ws.max_row, min_col=1, max_col=len(plumbing_headers)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row_idx in range(2, plumbing_ws.max_row + 1):
        val = plumbing_ws.cell(row=row_idx, column=2).value
        if val == "TOTAL REQUIRED":
            for col_idx in range(1, len(plumbing_headers) + 1):
                cell = plumbing_ws.cell(row=row_idx, column=col_idx)
                cell.font = Font(bold=True)
                cell.fill = accent_fill
        elif val == "ACCUMULATED (raw)":
            for col_idx in range(1, len(plumbing_headers) + 1):
                cell = plumbing_ws.cell(row=row_idx, column=col_idx)
                cell.font = Font(italic=True)
                cell.fill = PatternFill("solid", fgColor="F2F2F2")

    for col_idx, w in {1: 6, 2: 18, 3: 40, 4: 12, 5: 12, 6: 12, 7: 10, 8: 10,
                       9: 10, 10: 10, 11: 16, 12: 18, 13: 16, 14: 44, 15: 28}.items():
        plumbing_ws.column_dimensions[chr(64 + col_idx)].width = w

    # Code Basis sheet
    basis = wb.create_sheet("Code Basis")
    basis.append(["Standard", standard_name])
    basis.append(["Module", "Occupant Load + Plumbing Fixture Count Generator"])
    basis.append(["Occupant Load Method", "Occupant load = ceiling(room area / occupant load factor) per IBC Section 1004"])
    basis.append(["Plumbing Fixture Table", "IBC 2021 Table B2902.1"])
    basis.append(["Plumbing Fixture Method", "Raw fractional fixtures accumulated across categories; final total rounded up (ceiling)."])
    basis.append(["Separate Facilities", "IBC 2021 Section B2902.2"])
    basis.append(["Sex Distribution", f"Male ratio = {male_ratio:.0%}; female ratio = {1 - male_ratio:.0%}"])
    basis.append(["Note", "Results are an aid and require professional review."])
    for row in basis.iter_rows(min_row=1, max_row=basis.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = border
    for cell in basis[1]:
        cell.font = Font(bold=True)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------

st.markdown("## CodeCalc")
st.caption("Building code calculations for drawing-ready submittals")

with st.sidebar:
    st.markdown("### Settings")
    standard_name = st.selectbox("Occupant Load Standard", options=list(OCCUPANT_LOAD_STANDARDS.keys()), index=0)
    male_ratio = st.slider("Male occupant ratio", min_value=0.0, max_value=1.0, value=0.5, step=0.05,
                            help="Used to split total occupants into male/female for fixture calculations")

function_options = [format_function_label(item) for item in get_standard_table(standard_name)]
plumbing_options = [format_plumbing_category_label(item) for item in get_plumbing_category_table()]

# ── Occupant Load Schedule ─────────────────────────────────────────────────

st.markdown("### Occupant Load Schedule")
st.caption("IBC Section 1004 \u2014 enter rooms to calculate occupant loads")

if "rooms_df" not in st.session_state:
    st.session_state.rooms_df = starter_rows()

editable_df = st.data_editor(
    st.session_state.rooms_df,
    num_rows="dynamic",
    use_container_width=True,
    column_config={
        "Story": st.column_config.TextColumn("Story", required=True),
        "Room Number": st.column_config.TextColumn("Room No.", required=True),
        "Room Name": st.column_config.TextColumn("Room Name", required=True),
        "Room Area": st.column_config.NumberColumn("Area (sf)", min_value=0.0, step=1.0, format="%.0f", required=True),
        "Room Function": st.column_config.SelectboxColumn("Function of Space", options=function_options, required=True, width="large"),
    },
    hide_index=True,
    key="rooms_editor",
)
st.session_state.rooms_df = editable_df
schedule_df = enrich_rows(editable_df, standard_name)

if schedule_df.empty:
    st.info("Add at least one valid room row above.")
else:
    display_df = build_schedule_lines(schedule_df)
    st.dataframe(display_df, use_container_width=True, hide_index=True)

# ── Plumbing Fixture Count Analysis ───────────────────────────────────────

st.markdown("---")
st.markdown("### Plumbing Fixture Count Analysis")
st.caption("IBC 2021 Table B2902.1 \u2014 select categories and assign occupant counts")

if "plumbing_df" not in st.session_state:
    st.session_state.plumbing_df = plumbing_starter_rows()

plumbing_editable = st.data_editor(
    st.session_state.plumbing_df,
    num_rows="dynamic",
    use_container_width=True,
    column_config={
        "Plumbing Category": st.column_config.SelectboxColumn(
            "Table B2902.1 Classification",
            options=plumbing_options,
            required=True,
            width="large",
        ),
        "Occupants": st.column_config.NumberColumn("Occupants", min_value=0, step=1, required=True),
    },
    hide_index=True,
    key="plumbing_editor",
)
st.session_state.plumbing_df = plumbing_editable
plumbing_analysis = build_plumbing_analysis(plumbing_editable, male_ratio)

if not plumbing_analysis:
    st.info("Add at least one plumbing category row with occupants > 0.")
else:
    sums = sum_fixtures(plumbing_analysis)
    required = ceil_fixtures(sums)

    # Per-category detail blocks
    for item in plumbing_analysis:
        with st.container():
            st.markdown(
                f'<div class="plumbing-block">'
                f'<strong style="font-size:0.95rem;">{item["classification_no"]} \u2014 {item["classification"]}</strong>'
                f'<br><span style="color:#5a6577;font-size:0.82rem;">{item["description"]}</span>'
                f'<br><span style="font-size:0.8rem;">Occupants: <strong>{item["occupants"]}</strong> '
                f'(M: {item["male_occ"]}, F: {item["female_occ"]})</span>'
                f'</div>',
                unsafe_allow_html=True,
            )

            c1, c2, c3, c4, c5, c6, c7 = st.columns(7)
            c1.metric("WC (M)", fmt_frac(item["wc_male"]), help=format_ratio_text(item["wc_male_ratio"]))
            c2.metric("WC (F)", fmt_frac(item["wc_female"]), help=format_ratio_text(item["wc_female_ratio"]))
            c3.metric("Lav (M)", fmt_frac(item["lav_male"]), help=format_ratio_text(item["lav_male_ratio"]))
            c4.metric("Lav (F)", fmt_frac(item["lav_female"]), help=format_ratio_text(item["lav_female_ratio"]))
            c5.metric("Bath/Shower", fmt_frac(item["bath"]), help=format_ratio_text(item["bath_ratio"]))
            c6.metric("Drinking Ftn", fmt_frac(item["drinking"]), help=format_ratio_text(item["drinking_ratio"]))
            c7.metric("Service Sink", fmt_frac(item["service_sink"]), help=format_ratio_text(item["service_sink_ratio"]))

            sep = item["separate_facilities"]
            if sep.startswith("Required"):
                st.markdown(f'<div class="sep-required"><strong>B2902.2:</strong> {sep}</div>', unsafe_allow_html=True)
            else:
                st.markdown(f'<div class="sep-not-required"><strong>B2902.2:</strong> {sep}</div>', unsafe_allow_html=True)

            st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)

    # Accumulated raw totals
    if len(plumbing_analysis) > 1:
        st.markdown(
            '<div style="background:#f2f2f2;border-radius:6px;padding:0.6rem 1rem;margin-bottom:0.6rem;">'
            '<span style="font-size:0.82rem;font-weight:600;color:#5a6577;text-transform:uppercase;">Accumulated (fractional)</span>'
            '</div>',
            unsafe_allow_html=True,
        )
        a1, a2, a3, a4, a5, a6, a7 = st.columns(7)
        a1.metric("WC (M)", fmt_frac(sums["wc_male"]))
        a2.metric("WC (F)", fmt_frac(sums["wc_female"]))
        a3.metric("Lav (M)", fmt_frac(sums["lav_male"]))
        a4.metric("Lav (F)", fmt_frac(sums["lav_female"]))
        a5.metric("Bath/Shower", fmt_frac(sums["bath"]))
        a6.metric("Drinking Ftn", fmt_frac(sums["drinking"]))
        a7.metric("Service Sink", fmt_frac(sums["service_sink"]))

    # Final required (ceiling)
    st.markdown("#### Total Required")
    with st.container():
        st.markdown('<div class="totals-card">', unsafe_allow_html=True)
        t1, t2, t3, t4, t5, t6, t7 = st.columns(7)
        t1.metric("WC (M)", required["wc_male"])
        t2.metric("WC (F)", required["wc_female"])
        t3.metric("Lav (M)", required["lav_male"])
        t4.metric("Lav (F)", required["lav_female"])
        t5.metric("Bath/Shower", required["bath"] if required["bath"] > 0 else "\u2014")
        t6.metric("Drinking Ftn", required["drinking"])
        t7.metric("Service Sink", required["service_sink"] if required["service_sink"] > 0 else "\u2014")
        st.markdown('</div>', unsafe_allow_html=True)

    st.caption("Fractional fixture values accumulated across all categories; final total rounded up per code. "
               "Exceptions 5 & 6 of B2902.2 (single-user toilet rooms; rooms designed for both sexes) "
               "are design-based and must be evaluated by the design professional.")

# ── Excel download ────────────────────────────────────────────────────────

if not schedule_df.empty or plumbing_analysis:
    st.markdown("")
    st.download_button(
        label="\U0001F4E5  Download Excel",
        data=build_workbook(schedule_df, standard_name, plumbing_analysis, male_ratio),
        file_name="occupant_load_and_plumbing_fixture_analysis.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

st.markdown("---")
st.caption("Occupant load per IBC Section 1004 \u00b7 Plumbing fixtures per IBC 2021 Table B2902.1 \u00b7 Separate facilities per B2902.2")
